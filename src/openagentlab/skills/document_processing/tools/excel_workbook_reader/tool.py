from datetime import date, datetime
from pathlib import Path
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from openagentlab.skills.tool import BaseTool

from .schemas import (
    ExcelSheetInfo,
    ExcelWorkbookReaderInput,
    ExcelWorkbookReaderOutput,
)


class ExcelWorkbookReaderError(ValueError):
    pass


class ExcelWorkbookReaderTool(BaseTool):
    def __init__(self) -> None:
        super().__init__(
            name="excel_workbook_reader",
            description="Inspect local .xlsx workbook sheets and basic metadata.",
            capability="document.read.excel.workbook",
        )

    def execute(
        self,
        tool_input: ExcelWorkbookReaderInput,
    ) -> ExcelWorkbookReaderOutput:
        path = self._validate_path(tool_input.path)

        try:
            workbook = load_workbook(path, read_only=True, data_only=False)
        except (BadZipFile, InvalidFileException) as exc:
            msg = f"Could not read Excel workbook: {path}"
            raise ExcelWorkbookReaderError(msg) from exc

        try:
            sheets = tuple(
                ExcelSheetInfo(
                    name=worksheet.title,
                    index=index,
                    max_row=worksheet.max_row,
                    max_column=worksheet.max_column,
                )
                for index, worksheet in enumerate(workbook.worksheets, start=1)
            )

            return ExcelWorkbookReaderOutput(
                path=path,
                sheet_count=len(sheets),
                sheets=sheets,
                metadata=self._normalize_metadata(workbook.properties),
            )
        finally:
            workbook.close()

    def _validate_path(self, path: Path) -> Path:
        normalized_path = path.expanduser()

        if not normalized_path.exists():
            msg = f"Excel workbook does not exist: {normalized_path}"
            raise FileNotFoundError(msg)

        if not normalized_path.is_file():
            msg = f"Excel workbook input is not a file: {normalized_path}"
            raise IsADirectoryError(msg)

        if normalized_path.suffix.lower() != ".xlsx":
            msg = f"Excel workbook input must use a .xlsx extension: {normalized_path}"
            raise ValueError(msg)

        return normalized_path

    def _normalize_metadata(self, properties: object) -> dict[str, str | None]:
        metadata_fields = (
            "title",
            "subject",
            "creator",
            "description",
            "created",
            "modified",
        )

        return {
            field: self._serialize_metadata_value(getattr(properties, field, None))
            for field in metadata_fields
        }

    def _serialize_metadata_value(self, value: object) -> str | None:
        if value is None:
            return None

        if isinstance(value, datetime | date):
            return value.isoformat()

        return str(value)
