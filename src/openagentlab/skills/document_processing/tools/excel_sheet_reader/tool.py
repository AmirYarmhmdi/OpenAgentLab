"""File guide.

- Use: Implements the deterministic tool that reads one worksheet from an Excel
  workbook.
- Usage: Import ExcelSheetReaderError, and ExcelSheetReaderTool from
  openagentlab.skills.document_processing.tools.excel_sheet_reader.tool.
- Duties: Defines ExcelSheetReaderError, and ExcelSheetReaderTool and related helper
  logic.
- Depends on: Project modules: openagentlab.skills.document_processing.tools.excel_s
  heet_reader.schemas, and openagentlab.skills.tool.
"""

from datetime import date, datetime
from pathlib import Path
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from openagentlab.skills.tool import BaseTool

from .schemas import ExcelCellValue, ExcelSheetReaderInput, ExcelSheetReaderOutput


class ExcelSheetReaderError(ValueError):
    pass


class ExcelSheetReaderTool(BaseTool):
    def __init__(self) -> None:
        super().__init__(
            name="excel_sheet_reader",
            description=(
                "Extract worksheet rows from a named sheet in a local .xlsx file."
            ),
            capability="document.read.excel.sheet",
        )

    def execute(self, tool_input: ExcelSheetReaderInput) -> ExcelSheetReaderOutput:
        path = self._validate_path(tool_input.path)
        self._validate_max_rows(tool_input.max_rows)

        try:
            workbook = load_workbook(path, read_only=True, data_only=False)
        except (BadZipFile, InvalidFileException) as exc:
            msg = f"Could not read Excel workbook: {path}"
            raise ExcelSheetReaderError(msg) from exc

        try:
            if tool_input.sheet_name not in workbook.sheetnames:
                msg = f"Worksheet does not exist: {tool_input.sheet_name}"
                raise ExcelSheetReaderError(msg)

            worksheet = workbook[tool_input.sheet_name]
            returned_row_count = self._returned_row_count(
                worksheet.max_row,
                tool_input.max_rows,
            )
            rows = tuple(
                tuple(self._serialize_cell_value(value) for value in row)
                for row in worksheet.iter_rows(
                    min_row=1,
                    max_row=returned_row_count,
                    max_col=worksheet.max_column,
                    values_only=True,
                )
            )

            return ExcelSheetReaderOutput(
                path=path,
                sheet_name=worksheet.title,
                row_count=worksheet.max_row,
                column_count=worksheet.max_column,
                rows=rows,
                truncated=tool_input.max_rows is not None
                and tool_input.max_rows < worksheet.max_row,
            )
        finally:
            workbook.close()

    def _validate_path(self, path: Path) -> Path:
        normalized_path = path.expanduser()

        if not normalized_path.exists():
            msg = f"Excel workbook does not exist: {normalized_path}"
            raise FileNotFoundError(msg)

        if not normalized_path.is_file():
            msg = f"Excel workbook input is not a file: {normalized_path}"
            raise IsADirectoryError(msg)

        if normalized_path.suffix.lower() != ".xlsx":
            msg = f"Excel workbook input must use a .xlsx extension: {normalized_path}"
            raise ValueError(msg)

        return normalized_path

    def _validate_max_rows(self, max_rows: int | None) -> None:
        if max_rows is not None and max_rows < 1:
            msg = "max_rows must be positive"
            raise ValueError(msg)

    def _returned_row_count(self, row_count: int, max_rows: int | None) -> int:
        if max_rows is None:
            return row_count

        return min(row_count, max_rows)

    def _serialize_cell_value(self, value: object) -> ExcelCellValue:
        if value is None or isinstance(value, str | int | float | bool):
            return value

        if isinstance(value, datetime | date):
            return value.isoformat()

        return str(value)
