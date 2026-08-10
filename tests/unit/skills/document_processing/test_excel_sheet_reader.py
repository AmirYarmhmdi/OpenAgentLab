from datetime import date, datetime
from pathlib import Path

import pytest
from openpyxl import Workbook
from pydantic import ValidationError

from openagentlab.skills.document_processing import DocumentProcessingSkill
from openagentlab.skills.document_processing.tools.excel_sheet_reader import (
    ExcelSheetReaderError,
    ExcelSheetReaderInput,
    ExcelSheetReaderTool,
)


def write_sheet_reader_workbook(path: Path) -> None:
    workbook = Workbook()

    data_sheet = workbook.active
    data_sheet.title = "Data"
    data_sheet.append(["label", "amount", "active", "when"])
    data_sheet.append(["alpha", 10, True, datetime(2026, 8, 8, 9, 30, 0)])
    data_sheet.append(["beta", 12.5, False, date(2026, 8, 9)])
    data_sheet.append(["empty-middle", None, True, "tail"])

    formula_sheet = workbook.create_sheet("Formula")
    formula_sheet["A1"] = 1
    formula_sheet["A2"] = 2
    formula_sheet["A3"] = "=SUM(A1:A2)"

    other_sheet = workbook.create_sheet("Other")
    other_sheet.append(["other sheet"])

    workbook.save(path)
    workbook.close()


def test_excel_sheet_reader_tool_exposes_metadata() -> None:
    tool = ExcelSheetReaderTool()

    assert tool.name == "excel_sheet_reader"
    assert tool.capability == "document.read.excel.sheet"
    assert "worksheet" in tool.description


def test_excel_sheet_reader_reads_named_worksheet(tmp_path: Path) -> None:
    path = tmp_path / "sheet.xlsx"
    write_sheet_reader_workbook(path)

    result = ExcelSheetReaderTool().execute(
        ExcelSheetReaderInput(path=path, sheet_name="Data")
    )

    assert result.path == path
    assert result.sheet_name == "Data"
    assert result.row_count == 4
    assert result.column_count == 4
    assert result.truncated is False


def test_excel_sheet_reader_selects_exact_sheet_among_multiple_sheets(
    tmp_path: Path,
) -> None:
    path = tmp_path / "multi-sheet.xlsx"
    write_sheet_reader_workbook(path)

    result = ExcelSheetReaderTool().execute(
        ExcelSheetReaderInput(path=path, sheet_name="Other")
    )

    assert result.rows == (("other sheet",),)


def test_excel_sheet_reader_preserves_row_and_cell_order(tmp_path: Path) -> None:
    path = tmp_path / "ordered.xlsx"
    write_sheet_reader_workbook(path)

    result = ExcelSheetReaderTool().execute(
        ExcelSheetReaderInput(path=path, sheet_name="Data")
    )

    assert result.rows[0] == ("label", "amount", "active", "when")
    assert result.rows[1][:3] == ("alpha", 10, True)
    assert result.rows[2][:3] == ("beta", 12.5, False)


def test_excel_sheet_reader_preserves_empty_cells(tmp_path: Path) -> None:
    path = tmp_path / "empty-cell.xlsx"
    write_sheet_reader_workbook(path)

    result = ExcelSheetReaderTool().execute(
        ExcelSheetReaderInput(path=path, sheet_name="Data")
    )

    assert result.rows[3] == ("empty-middle", None, True, "tail")


def test_excel_sheet_reader_normalizes_date_and_datetime_values(
    tmp_path: Path,
) -> None:
    path = tmp_path / "dates.xlsx"
    write_sheet_reader_workbook(path)

    result = ExcelSheetReaderTool().execute(
        ExcelSheetReaderInput(path=path, sheet_name="Data")
    )

    assert result.rows[1][3] == "2026-08-08T09:30:00"
    assert result.rows[2][3] == "2026-08-09T00:00:00"


def test_excel_sheet_reader_returns_formula_expressions(tmp_path: Path) -> None:
    path = tmp_path / "formula.xlsx"
    write_sheet_reader_workbook(path)

    result = ExcelSheetReaderTool().execute(
        ExcelSheetReaderInput(path=path, sheet_name="Formula")
    )

    assert result.rows == ((1,), (2,), ("=SUM(A1:A2)",))


def test_excel_sheet_reader_requires_exact_sheet_name(tmp_path: Path) -> None:
    path = tmp_path / "exact.xlsx"
    write_sheet_reader_workbook(path)

    with pytest.raises(ExcelSheetReaderError, match="Worksheet does not exist"):
        ExcelSheetReaderTool().execute(
            ExcelSheetReaderInput(path=path, sheet_name="data")
        )


def test_excel_sheet_reader_rejects_missing_worksheet(tmp_path: Path) -> None:
    path = tmp_path / "missing-sheet.xlsx"
    write_sheet_reader_workbook(path)

    with pytest.raises(ExcelSheetReaderError, match="Missing"):
        ExcelSheetReaderTool().execute(
            ExcelSheetReaderInput(path=path, sheet_name="Missing")
        )


def test_excel_sheet_reader_rejects_missing_file(tmp_path: Path) -> None:
    path = tmp_path / "missing.xlsx"

    with pytest.raises(FileNotFoundError, match="missing.xlsx"):
        ExcelSheetReaderTool().execute(
            ExcelSheetReaderInput(path=path, sheet_name="Data")
        )


def test_excel_sheet_reader_rejects_directory_input(tmp_path: Path) -> None:
    with pytest.raises(IsADirectoryError, match="not a file"):
        ExcelSheetReaderTool().execute(
            ExcelSheetReaderInput(path=tmp_path, sheet_name="Data")
        )


def test_excel_sheet_reader_rejects_unsupported_extension(tmp_path: Path) -> None:
    path = tmp_path / "sheet.xls"
    path.write_bytes(b"not supported")

    with pytest.raises(ValueError, match=".xlsx extension"):
        ExcelSheetReaderTool().execute(
            ExcelSheetReaderInput(path=path, sheet_name="Data")
        )


def test_excel_sheet_reader_rejects_malformed_workbook(tmp_path: Path) -> None:
    path = tmp_path / "broken.xlsx"
    path.write_bytes(b"not a real workbook")

    with pytest.raises(ExcelSheetReaderError, match="Could not read Excel workbook"):
        ExcelSheetReaderTool().execute(
            ExcelSheetReaderInput(path=path, sheet_name="Data")
        )


def test_excel_sheet_reader_validates_max_rows(tmp_path: Path) -> None:
    path = tmp_path / "max-rows.xlsx"
    write_sheet_reader_workbook(path)

    with pytest.raises(ValidationError):
        ExcelSheetReaderInput(path=path, sheet_name="Data", max_rows=0)


def test_excel_sheet_reader_limits_returned_rows_and_marks_truncated(
    tmp_path: Path,
) -> None:
    path = tmp_path / "limited.xlsx"
    write_sheet_reader_workbook(path)

    result = ExcelSheetReaderTool().execute(
        ExcelSheetReaderInput(path=path, sheet_name="Data", max_rows=2)
    )

    assert result.row_count == 4
    assert len(result.rows) == 2
    assert result.rows == (
        ("label", "amount", "active", "when"),
        ("alpha", 10, True, "2026-08-08T09:30:00"),
    )
    assert result.truncated is True


def test_excel_sheet_reader_max_rows_without_truncation(tmp_path: Path) -> None:
    path = tmp_path / "not-truncated.xlsx"
    write_sheet_reader_workbook(path)

    result = ExcelSheetReaderTool().execute(
        ExcelSheetReaderInput(path=path, sheet_name="Data", max_rows=10)
    )

    assert len(result.rows) == 4
    assert result.truncated is False


def test_document_processing_skill_exposes_excel_sheet_reader_tool() -> None:
    skill = DocumentProcessingSkill()

    assert "document.read.pdf" in skill.executable_capabilities
    assert "document.read.excel.workbook" in skill.executable_capabilities
    assert "document.read.excel.sheet" in skill.executable_capabilities
    assert "document.read.csv" in skill.executable_capabilities
    assert "document.read.text" in skill.executable_capabilities
    assert "document.read.json" in skill.executable_capabilities
    assert "document.read.docx" in skill.executable_capabilities
    assert skill.get_tool("excel_sheet_reader") is not None
