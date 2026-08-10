"""File guide.

- Use: Contains unit tests for excel workbook reader behavior.
- Usage: Run this file with pytest when checking related behavior.
- Duties: Builds test data, calls the public API, and checks expected results.
- Depends on: Project modules: openagentlab.skills.document_processing, and
  openagentlab.skills.document_processing.tools.excel_workbook_reader.
"""

from datetime import datetime
from pathlib import Path

import pytest
from openpyxl import Workbook

from openagentlab.skills.document_processing import DocumentProcessingSkill
from openagentlab.skills.document_processing.tools.excel_workbook_reader import (
    ExcelWorkbookReaderError,
    ExcelWorkbookReaderInput,
    ExcelWorkbookReaderTool,
)


def write_workbook(path: Path) -> None:
    workbook = Workbook()

    first_sheet = workbook.active
    first_sheet.title = "Summary"
    first_sheet.append(["name", "value"])
    first_sheet.append(["alpha", 1])

    details_sheet = workbook.create_sheet("Details")
    details_sheet["A1"] = "id"
    details_sheet["B1"] = "status"
    details_sheet["C4"] = "done"

    workbook.properties.title = "Quarterly Workbook"
    workbook.properties.subject = "Workbook inspection"
    workbook.properties.creator = "OpenAgentLab"
    workbook.properties.description = "Small deterministic workbook fixture."
    workbook.properties.created = datetime(2026, 8, 8, 9, 30, 0)
    workbook.properties.modified = datetime(2026, 8, 8, 10, 45, 0)

    workbook.save(path)
    workbook.close()


def test_excel_workbook_reader_tool_exposes_metadata() -> None:
    tool = ExcelWorkbookReaderTool()

    assert tool.name == "excel_workbook_reader"
    assert "workbook" in tool.description
    assert tool.capability == "document.read.excel.workbook"


def test_excel_workbook_reader_reads_xlsx_workbook(tmp_path: Path) -> None:
    path = tmp_path / "workbook.xlsx"
    write_workbook(path)

    result = ExcelWorkbookReaderTool().execute(ExcelWorkbookReaderInput(path=path))

    assert result.path == path
    assert result.sheet_count == 2


def test_excel_workbook_reader_preserves_sheet_order_names_and_indices(
    tmp_path: Path,
) -> None:
    path = tmp_path / "ordered.xlsx"
    write_workbook(path)

    result = ExcelWorkbookReaderTool().execute(ExcelWorkbookReaderInput(path=path))

    assert [(sheet.index, sheet.name) for sheet in result.sheets] == [
        (1, "Summary"),
        (2, "Details"),
    ]


def test_excel_workbook_reader_reports_worksheet_dimensions(tmp_path: Path) -> None:
    path = tmp_path / "dimensions.xlsx"
    write_workbook(path)

    result = ExcelWorkbookReaderTool().execute(ExcelWorkbookReaderInput(path=path))

    assert result.sheets[0].max_row == 2
    assert result.sheets[0].max_column == 2
    assert result.sheets[1].max_row == 4
    assert result.sheets[1].max_column == 3


def test_excel_workbook_reader_returns_basic_workbook_metadata(
    tmp_path: Path,
) -> None:
    path = tmp_path / "metadata.xlsx"
    write_workbook(path)

    result = ExcelWorkbookReaderTool().execute(ExcelWorkbookReaderInput(path=path))

    assert result.metadata["title"] == "Quarterly Workbook"
    assert result.metadata["subject"] == "Workbook inspection"
    assert result.metadata["creator"] == "OpenAgentLab"
    assert result.metadata["description"] == "Small deterministic workbook fixture."
    assert result.metadata["created"] == "2026-08-08T09:30:00"
    assert result.metadata["modified"] is not None


def test_excel_workbook_reader_rejects_missing_file(tmp_path: Path) -> None:
    path = tmp_path / "missing.xlsx"

    with pytest.raises(FileNotFoundError, match="missing.xlsx"):
        ExcelWorkbookReaderTool().execute(ExcelWorkbookReaderInput(path=path))


def test_excel_workbook_reader_rejects_directory_input(tmp_path: Path) -> None:
    with pytest.raises(IsADirectoryError, match="not a file"):
        ExcelWorkbookReaderTool().execute(ExcelWorkbookReaderInput(path=tmp_path))


def test_excel_workbook_reader_rejects_unsupported_extension(tmp_path: Path) -> None:
    path = tmp_path / "workbook.xls"
    path.write_bytes(b"not supported")

    with pytest.raises(ValueError, match=".xlsx extension"):
        ExcelWorkbookReaderTool().execute(ExcelWorkbookReaderInput(path=path))


def test_excel_workbook_reader_rejects_malformed_workbook(tmp_path: Path) -> None:
    path = tmp_path / "broken.xlsx"
    path.write_bytes(b"not a real workbook")

    with pytest.raises(ExcelWorkbookReaderError, match="Could not read Excel workbook"):
        ExcelWorkbookReaderTool().execute(ExcelWorkbookReaderInput(path=path))


def test_document_processing_skill_exposes_excel_workbook_reader_tool() -> None:
    skill = DocumentProcessingSkill()

    assert "document.read.pdf" in skill.executable_capabilities
    assert "document.read.excel.workbook" in skill.executable_capabilities
    assert "document.read.excel.sheet" in skill.executable_capabilities
    assert "document.read.csv" in skill.executable_capabilities
    assert "document.read.text" in skill.executable_capabilities
    assert "document.read.json" in skill.executable_capabilities
    assert "document.read.docx" in skill.executable_capabilities
    assert skill.get_tool("excel_workbook_reader") is not None
